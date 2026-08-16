import asyncio
import os
import re
from collections import Counter
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
URL_TARGET = "http://books.toscrape.com/"
NAMA_FILE_OUTPUT = os.path.join(BASE_DIR, "daftar_buku_premium_lengkap.xlsx")
HEADLESS_MODE = False  # ganti True kalau mau proses lebih cepat & tanpa tampilan browser

RATING_DICT = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}

# Theme colors
WARNA_HEADER = "1F4E78"      # navy premium
WARNA_ZEBRA = "F2F6FC"       # biru muda buat baris selang-seling
WARNA_BINTANG = "FFC000"     # emas/kuning
WARNA_STOK_AMAN = "C6E0B4"   # hijau
WARNA_STOK_WASPADA = "FFE699"  # kuning
WARNA_STOK_KRITIS = "F8CBAD"   # merah muda


async def main():
    async with async_playwright() as p:
        print("Deep Scraper Bot is preparing to launch the browser...")
        browser = await p.chromium.launch(headless=HEADLESS_MODE)
        page = await browser.new_page()

        print(f"Navigating to: {URL_TARGET}")
        await page.goto(URL_TARGET, wait_until="load", timeout=30000)
        await page.wait_for_timeout(2000)

        print("Attempting to retrieve the initial book product list...")
        try:
            await page.wait_for_selector("article.product_pod", timeout=10000)
            buku_elements = await page.locator("article.product_pod").all()
            print(f"🎉 Success! Found {len(buku_elements)} initial book products.")

            # Store initial data from the listing page
            daftar_buku_data = []
            for buku in buku_elements:
                link_element = buku.locator("h3 a")
                judul = await link_element.get_attribute("title")
                href = await link_element.get_attribute("href")
                url_detail = f"{URL_TARGET}{href}" if href.startswith("catalogue/") else f"{URL_TARGET}catalogue/{href}"

                rating_classes = await buku.locator("p.star-rating").get_attribute("class")
                rating_text = rating_classes.replace("star-rating ", "").strip()

                daftar_buku_data.append({
                    "judul": judul,
                    "url_detail": url_detail,
                    "rating_angka": RATING_DICT.get(rating_text, 0),
                })

            # ============================================================
            # DEEP SCRAPING OF EACH BOOK DETAIL PAGE
            # ============================================================
            print("\n🕵️  Starting deep scraping of each product detail page...")
            new_page = await browser.new_page()
            semua_data = []

            for index, data in enumerate(daftar_buku_data, start=1):
                print(f"[{index}/{len(daftar_buku_data)}] Visiting: {data['judul'][:40]}...")
                try:
                    await new_page.goto(data["url_detail"], wait_until="load", timeout=15000)

                    # Category dari breadcrumb
                    kategori_buku = await new_page.locator("ul.breadcrumb li:nth-child(3) a").inner_text()

                    # Tabel informasi produk (UPC, harga, pajak, stok, review, dll)
                    info_rows = await new_page.locator("table.table.table-striped tr").all()
                    info = {}
                    for row in info_rows:
                        th_text = await row.locator("th").inner_text()
                        td_text = await row.locator("td").inner_text()
                        info[th_text.strip()] = td_text.strip()

                    upc = info.get("UPC", "N/A")
                    tipe_produk = info.get("Product Type", "N/A")
                    harga_excl = info.get("Price (excl. tax)", "N/A")
                    harga_incl = info.get("Price (incl. tax)", "N/A")
                    pajak = info.get("Tax", "N/A")
                    availability_raw = info.get("Availability", "Unknown")
                    jumlah_review = info.get("Number of reviews", "0")

                    # Extract the STOCK QUANTITY from text "In stock (22 available)"
                    match_stok = re.search(r"\((\d+)\s+available\)", availability_raw)
                    jumlah_stok = int(match_stok.group(1)) if match_stok else 0
                    status_stok = "In Stock" if "In stock" in availability_raw else "Out of Stock"

                    # Product description (truncated for Excel readability)
                    try:
                        deskripsi = await new_page.locator("#product_description + p").inner_text()
                        deskripsi_singkat = (deskripsi[:160] + "...") if len(deskripsi) > 160 else deskripsi
                    except Exception:
                        deskripsi_singkat = "No description available"

                except Exception as e:
                    kategori_buku = "General (Failed to Load)"
                    upc = tipe_produk = harga_excl = harga_incl = pajak = "N/A"
                    jumlah_stok = 0
                    status_stok = "Unknown"
                    jumlah_review = "0"
                    deskripsi_singkat = "Failed to load description"

                semua_data.append({
                    "judul": data["judul"],
                    "kategori": kategori_buku,
                    "upc": upc,
                    "tipe_produk": tipe_produk,
                    "harga_excl": harga_excl,
                    "harga_incl": harga_incl,
                    "pajak": pajak,
                    "rating_angka": data["rating_angka"],
                    "jumlah_review": jumlah_review,
                    "status_stok": status_stok,
                    "jumlah_stok": jumlah_stok,
                    "deskripsi": deskripsi_singkat,
                })

            await new_page.close()

            # ============================================================
            # SHEET 1: DETAILED BOOK CATALOG
            # ============================================================
            wb = Workbook()
            ws = wb.active
            ws.title = "Detailed Book Catalog"

            headers = [
                "No", "Book Title", "Category", "UPC", "Product Type",
                "Price (Excl. Tax)", "Price (Incl. Tax)", "Tax",
                "Rating", "Review Count", "Stock Status", "Stock Quantity", "Short Description"
            ]

            # Main title in row 1
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            judul_cell = ws.cell(row=1, column=1, value="🏆 DEEP SCRAPING REPORT — BOOKS TO SCRAPE")
            judul_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            judul_cell.fill = PatternFill(start_color=WARNA_HEADER, end_color=WARNA_HEADER, fill_type="solid")
            judul_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # Column headers in row 2
            header_fill = PatternFill(start_color=WARNA_HEADER, end_color=WARNA_HEADER, fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"),
            )

            for col_num, header_text in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_num, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[2].height = 22

            # Data starts from row 3
            baris = 3
            for i, d in enumerate(semua_data, start=1):
                bintang = "★" * d["rating_angka"] + "☆" * (5 - d["rating_angka"])
                row_values = [
                    i, d["judul"], d["kategori"], d["upc"], d["tipe_produk"],
                    d["harga_excl"], d["harga_incl"], d["pajak"],
                    bintang, d["jumlah_review"], d["status_stok"], d["jumlah_stok"], d["deskripsi"],
                ]
                for col_num, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=baris, column=col_num, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=(col_num == 13))
                    if col_num == 9:  # kolom Rating -> warna emas
                        cell.font = Font(color=WARNA_BINTANG, bold=True, size=12)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    if col_num == 1 or col_num == 12:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                # Zebra striping
                if i % 2 == 0:
                    for col_num in range(1, len(headers) + 1):
                        c = ws.cell(row=baris, column=col_num)
                        if col_num != 9:  # jangan timpa warna font bintang
                            c.fill = PatternFill(start_color=WARNA_ZEBRA, end_color=WARNA_ZEBRA, fill_type="solid")
                baris += 1

            baris_terakhir = baris - 1

            # Conditional formatting for the stock quantity column "Stock Quantity" (kolom ke-12 -> L)
            kolom_stok = get_column_letter(12)
            rentang_stok = f"{kolom_stok}3:{kolom_stok}{baris_terakhir}"
            ws.conditional_formatting.add(
                rentang_stok,
                CellIsRule(operator="lessThan", formula=["5"],
                           fill=PatternFill(start_color=WARNA_STOK_KRITIS, end_color=WARNA_STOK_KRITIS, fill_type="solid")),
            )
            ws.conditional_formatting.add(
                rentang_stok,
                CellIsRule(operator="between", formula=["5", "15"],
                           fill=PatternFill(start_color=WARNA_STOK_WASPADA, end_color=WARNA_STOK_WASPADA, fill_type="solid")),
            )
            ws.conditional_formatting.add(
                rentang_stok,
                CellIsRule(operator="greaterThan", formula=["15"],
                           fill=PatternFill(start_color=WARNA_STOK_AMAN, end_color=WARNA_STOK_AMAN, fill_type="solid")),
            )

            # Automatic column width
            lebar_khusus = {13: 45, 2: 35, 3: 18}  # deskripsi, judul, kategori dilebarin
            for col_num in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_num)
                if col_num in lebar_khusus:
                    ws.column_dimensions[col_letter].width = lebar_khusus[col_num]
                    continue
                max_len = max((len(str(ws.cell(row=r, column=col_num).value or "")) for r in range(2, baris_terakhir + 1)), default=10)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            ws.freeze_panes = "A3"
            ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{baris_terakhir}"

            # ============================================================
            # SHEET 2: DASHBOARD & STATISTICS (with charts)
            # ============================================================
            ws2 = wb.create_sheet("Dashboard & Statistik")

            ws2.merge_cells("A1:D1")
            t = ws2.cell(row=1, column=1, value="📊 BOOK CATALOG ANALYTICS DASHBOARD")
            t.font = Font(size=16, bold=True, color="FFFFFF")
            t.fill = PatternFill(start_color=WARNA_HEADER, end_color=WARNA_HEADER, fill_type="solid")
            t.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[1].height = 26

            # --- Summary statistics ---
            harga_float = []
            for d in semua_data:
                try:
                    harga_float.append(float(d["harga_incl"].replace("£", "").strip()))
                except ValueError:
                    pass
            total_buku = len(semua_data)
            total_stok = sum(d["jumlah_stok"] for d in semua_data)
            rata_harga = sum(harga_float) / len(harga_float) if harga_float else 0
            harga_min = min(harga_float) if harga_float else 0
            harga_max = max(harga_float) if harga_float else 0

            stat_label_font = Font(bold=True, size=11)
            stat_rows = [
                ("Total Books Scraped", total_buku),
                ("Total Available Stock", total_stok),
                ("Average Price", f"£{rata_harga:.2f}"),
                ("Lowest Price", f"£{harga_min:.2f}"),
                ("Highest Price", f"£{harga_max:.2f}"),
            ]
            r = 3
            for label, value in stat_rows:
                ws2.cell(row=r, column=1, value=label).font = stat_label_font
                ws2.cell(row=r, column=2, value=value)
                r += 1

            # --- Category distribution table (for bar chart) ---
            baris_kategori_awal = r + 2
            ws2.cell(row=baris_kategori_awal - 1, column=1, value="Distribusi Buku per Category").font = Font(bold=True, size=12)
            ws2.cell(row=baris_kategori_awal, column=1, value="Category").font = header_font
            ws2.cell(row=baris_kategori_awal, column=1).fill = header_fill
            ws2.cell(row=baris_kategori_awal, column=2, value="Count").font = header_font
            ws2.cell(row=baris_kategori_awal, column=2).fill = header_fill

            kategori_counter = Counter(d["kategori"] for d in semua_data)
            baris_kat = baris_kategori_awal + 1
            for kategori, jumlah in kategori_counter.most_common():
                ws2.cell(row=baris_kat, column=1, value=kategori)
                ws2.cell(row=baris_kat, column=2, value=jumlah)
                baris_kat += 1
            baris_kategori_akhir = baris_kat - 1

            chart_kategori = BarChart()
            chart_kategori.title = "Count Buku per Category"
            chart_kategori.style = 10
            chart_kategori.y_axis.title = "Count Buku"
            chart_kategori.x_axis.title = "Category"
            data_ref = Reference(ws2, min_col=2, min_row=baris_kategori_awal, max_row=baris_kategori_akhir)
            cat_ref = Reference(ws2, min_col=1, min_row=baris_kategori_awal + 1, max_row=baris_kategori_akhir)
            chart_kategori.add_data(data_ref, titles_from_data=True)
            chart_kategori.set_categories(cat_ref)
            chart_kategori.width = 18
            chart_kategori.height = 10
            ws2.add_chart(chart_kategori, f"D{baris_kategori_awal}")

            # --- Rating distribution table (for pie chart) ---
            baris_rating_awal = baris_kategori_akhir + 3
            ws2.cell(row=baris_rating_awal - 1, column=1, value="Rating Distribution").font = Font(bold=True, size=12)
            ws2.cell(row=baris_rating_awal, column=1, value="Rating").font = header_font
            ws2.cell(row=baris_rating_awal, column=1).fill = header_fill
            ws2.cell(row=baris_rating_awal, column=2, value="Count").font = header_font
            ws2.cell(row=baris_rating_awal, column=2).fill = header_fill

            rating_counter = Counter(d["rating_angka"] for d in semua_data)
            baris_rat = baris_rating_awal + 1
            for bintang_n in sorted(rating_counter.keys()):
                label = "★" * bintang_n + "☆" * (5 - bintang_n)
                ws2.cell(row=baris_rat, column=1, value=label)
                ws2.cell(row=baris_rat, column=2, value=rating_counter[bintang_n])
                baris_rat += 1
            baris_rating_akhir = baris_rat - 1

            chart_rating = PieChart()
            chart_rating.title = "Rating Distribution Buku"
            chart_rating.style = 10
            data_ref2 = Reference(ws2, min_col=2, min_row=baris_rating_awal, max_row=baris_rating_akhir)
            cat_ref2 = Reference(ws2, min_col=1, min_row=baris_rating_awal + 1, max_row=baris_rating_akhir)
            chart_rating.add_data(data_ref2, titles_from_data=True)
            chart_rating.set_categories(cat_ref2)
            chart_rating.width = 14
            chart_rating.height = 10
            ws2.add_chart(chart_rating, f"D{baris_rating_awal}")

            # --- Top 10 books with lowest stock (for bar chart) ---
            baris_stok_awal = baris_rating_akhir + 3
            ws2.cell(row=baris_stok_awal - 1, column=1, value="10 Books with Lowest Stock").font = Font(bold=True, size=12)
            ws2.cell(row=baris_stok_awal, column=1, value="Book Title").font = header_font
            ws2.cell(row=baris_stok_awal, column=1).fill = header_fill
            ws2.cell(row=baris_stok_awal, column=2, value="Stock Quantity").font = header_font
            ws2.cell(row=baris_stok_awal, column=2).fill = header_fill

            stok_terendah = sorted(semua_data, key=lambda d: d["jumlah_stok"])[:10]
            baris_s = baris_stok_awal + 1
            for d in stok_terendah:
                judul_pendek = d["judul"] if len(d["judul"]) <= 35 else d["judul"][:32] + "..."
                ws2.cell(row=baris_s, column=1, value=judul_pendek)
                ws2.cell(row=baris_s, column=2, value=d["jumlah_stok"])
                baris_s += 1
            baris_stok_akhir = baris_s - 1

            chart_stok = BarChart()
            chart_stok.type = "bar"  # horizontal, lebih enak dibaca untuk nama judul panjang
            chart_stok.title = "Lowest Stock (Top 10)"
            chart_stok.style = 11
            data_ref3 = Reference(ws2, min_col=2, min_row=baris_stok_awal, max_row=baris_stok_akhir)
            cat_ref3 = Reference(ws2, min_col=1, min_row=baris_stok_awal + 1, max_row=baris_stok_akhir)
            chart_stok.add_data(data_ref3, titles_from_data=True)
            chart_stok.set_categories(cat_ref3)
            chart_stok.width = 18
            chart_stok.height = 12
            ws2.add_chart(chart_stok, f"D{baris_stok_awal}")

            ws2.column_dimensions["A"].width = 38
            ws2.column_dimensions["B"].width = 16

            # ============================================================
            # SAVE FILE
            # ============================================================
            wb.save(NAMA_FILE_OUTPUT)
            print(f"\n🚀 SUCCESS! Excel report is ready at:\n{NAMA_FILE_OUTPUT}\n")

        except Exception as e:
            print(f"\nFailed to extract data. Error: {e}\n")

        await browser.close()
        print("Bot has finished its task.")


asyncio.run(main())
